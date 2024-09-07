import psutil
import time
import openpyxl
from openpyxl import Workbook, load_workbook
import socket
import ctypes
from ctypes import wintypes
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
import shutil
import os

def bytes_to_mb(bytes):
    """Convert bytes to megabytes"""
    return bytes / (1024 * 1024)

def get_network_traffic():
    """Get the current network traffic in megabytes"""
    net_io = psutil.net_io_counters()
    bytes_sent = net_io.bytes_sent
    bytes_received = net_io.bytes_recv
    return bytes_sent, bytes_received

def get_ip_address():
    """Get the current IP address"""
    s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
    s.settimeout(0)
    try:
        s.connect(('10.254.254.254', 1))
        ip_address = s.getsockname()[0]
    except Exception:
        ip_address = 'Desconocida'
    finally:
        s.close()
    return ip_address

def get_active_window_name():
    """Get the name of the currently active window"""
    user32 = ctypes.windll.user32
    kernel32 = ctypes.windll.kernel32

    hwnd = user32.GetForegroundWindow()
    pid = wintypes.DWORD()
    user32.GetWindowThreadProcessId(hwnd, ctypes.byref(pid))
    
    try:
        process = psutil.Process(pid.value)
        return process.name()
    except (psutil.NoSuchProcess, psutil.AccessDenied, psutil.ZombieProcess):
        return 'Desconocido'

def send_email(original_file_path, last_date):
    """Send an email with the specified file attached"""
    from_email = "sosahijas@gmail.com"
    to_email = "sosahijas@gmail.com"
    subject = "Informe de tráfico de red"
    body = "Adjunto el informe de tráfico de red."

    smtp_server = "smtp.gmail.com"
    smtp_port = 587
    password = "funb tvnf puit wrmh"  # Contraseña de aplicación

    ip_address = get_ip_address()
    new_file_name = f"{last_date},{ip_address}.xlsx"
    new_file_path = f"copia_{new_file_name}"
    shutil.copy(original_file_path, new_file_path)
    
    msg = MIMEMultipart()
    msg['From'] = from_email
    msg['To'] = to_email
    msg['Subject'] = subject
    msg.attach(MIMEText(body, 'plain', 'utf-8'))

    with open(new_file_path, "rb") as file:
        part = MIMEApplication(file.read(), Name=new_file_name)
    part['Content-Disposition'] = f'attachment; filename="{new_file_name}"'
    msg.attach(part)

    try:
        server = smtplib.SMTP(smtp_server, smtp_port)
        server.starttls()
        server.login(from_email, password)
        server.sendmail(from_email, to_email, msg.as_string())
        print("Correo electrónico enviado con éxito.")
    except Exception as e:
        print(f"Error al enviar el correo: {e}")
    finally:
        server.quit()

    remove_copied_files()

def create_new_excel_file(xlsx_file):
    """Create a new Excel file with headers"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Tráfico de Red"
    ws.append(['Fecha', 'Hora', 'Minuto', 'Segundo', 'Bytes Enviados (MB)', 'Bytes Recibidos (MB)', 'Total Enviado (MB)', 'Total Recibido (MB)', 'Total General (MB)', 'IP', 'Aplicación Activa'])
    wb.save(xlsx_file)

def remove_copied_files():
    """Remove files that start with 'copia_'"""
    for file in os.listdir('.'):
        if file.startswith('copia_') and file.endswith('.xlsx'):
            os.remove(file)
            print(f"Archivo eliminado: {file}")

def read_last_date_from_file(file_path):
    """Read the last date from the text file"""
    if os.path.exists(file_path):
        with open(file_path, 'r') as file:
            return file.read().strip()
    return ""

def write_current_date_to_file(file_path):
    """Write the current date to the text file"""
    current_date = time.strftime('%d-%m-%Y')
    with open(file_path, 'w') as file:
        file.write(current_date)
    print(f"Fecha actual escrita en el archivo: {current_date}")

def clear_excel_content(xlsx_file):
    """Clear content in Excel file from the second row onwards"""
    wb = load_workbook(xlsx_file)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.value = None
    wb.save(xlsx_file)
    print(f"Contenido del archivo {xlsx_file} limpiado desde la segunda fila.")

def main():
    interval = 1
    xlsx_file = 'trafico_red.xlsx'
    date_file = 'fecha.txt'
    
    # Crear el archivo Excel solo si no existe
    if not os.path.exists(xlsx_file):
        create_new_excel_file(xlsx_file)

    total_sent = 0
    total_received = 0
    last_date = time.strftime('%d-%m-%Y')

    ip_address = get_ip_address()
    print(f"IP de la máquina: {ip_address}")
    print("Monitoreando tráfico de red. Presiona Ctrl+C para detener.")

    last_recorded_date = read_last_date_from_file(date_file)

    if last_recorded_date != last_date:
        send_email(xlsx_file, last_recorded_date)
        write_current_date_to_file(date_file)

    try:
        while True:
            current_date = time.strftime('%d-%m-%Y')

            if current_date != last_date:
                send_email(xlsx_file, last_date)
                # Limpiar el contenido del Excel antes de comenzar a escribir nuevamente
                clear_excel_content(xlsx_file)
                last_date = current_date
                write_current_date_to_file(date_file)
                print(f"\nNuevo día detectado. Informe enviado por correo y archivo limpiado.")
            
            bytes_sent_before, bytes_received_before = get_network_traffic()
            time.sleep(interval)
            bytes_sent_after, bytes_received_after = get_network_traffic()
            
            sent_in_mb = bytes_to_mb(bytes_sent_after - bytes_sent_before)
            received_in_mb = bytes_to_mb(bytes_received_after - bytes_received_before)
            
            total_sent += sent_in_mb
            total_received += received_in_mb
            total_general = total_sent + total_received

            current_time = time.strftime('%Y-%m-%d %H:%M:%S')
            date, time_str = current_time.split(' ')
            hour, minute, second = time_str.split(':')

            active_app = get_active_window_name()
            print(f"Aplicación activa: {active_app}")

            wb = load_workbook(xlsx_file)
            ws = wb.active
            ws.append([date, hour, minute, second, f"{sent_in_mb:.2f}", f"{received_in_mb:.2f}", f"{total_sent:.2f}", f"{total_received:.2f}", f"{total_general:.2f}", ip_address, active_app])
            wb.save(xlsx_file)
            
            print(f"Enviado: {sent_in_mb:.2f} MB/s | Recibido: {received_in_mb:.2f} MB/s")

    except KeyboardInterrupt:
        print("\nMonitoreo detenido.")

if __name__ == "__main__":
    main()
